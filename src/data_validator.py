#coding:utf-8
import os
import openpyxl
workbook = openpyxl.load_workbook("C:\\Users\\allen.w\\OneDrive\\桌面\\元宵老公公\\男宿老公公\\非物質交換禮物配對表.xlsx")
worksheet = workbook["更正"]
domainlist=[]
for i in range(worksheet.max_row):
    personemail = worksheet.cell(row=i+1, column=4).value
    print(worksheet.cell(row=i+1, column=4).value)
    index = personemail.index("@")
    domain = personemail[index:]
    domainlist
    if domain not in domainlist:
        domainlist.append(domain)
print(domainlist)