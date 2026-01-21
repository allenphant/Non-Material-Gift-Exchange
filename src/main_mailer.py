# coding:utf-8
import os
import openpyxl
import yagmail
import time
from dotenv import load_dotenv  # 建議使用環境變數管理密碼

# 原始檔案有太多個人資料，這裡只保留必要的結構

# ================= 設定區 =================
# 請將這些資訊改為你的設定，或使用環境變數
SENDER_EMAIL = "your_email@gmail.com"  # 寄件者信箱
# SENDER_PASSWORD = Google應用程式管理密碼，請參考https://www.ibest.com.tw/news-detail/gmail-smtp/尋找自己的Google應用程式管理密碼
SENDER_PASSWORD = "your_google_app_password_here" # 如果只是本機執行，暫時填這裡，上傳前記得刪掉

EXCEL_PATH = "data/gift_exchange_list.xlsx" # Excel 檔案路徑
SHEET_NAME = "更正" # 工作表名稱
# =========================================

def send_gift_emails():
    # 檢查檔案是否存在
    if not os.path.exists(EXCEL_PATH):
        print(f"錯誤：找不到檔案 {EXCEL_PATH}")
        return

    print("正在讀取 Excel 檔案...")
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    worksheet = workbook[SHEET_NAME]
    max_row = worksheet.max_row
    
    # 初始化 Yagmail SMTP
    try:
        yag = yagmail.SMTP(SENDER_EMAIL, SENDER_PASSWORD)
        print("SMTP 連線成功！")
    except Exception as e:
        print(f"SMTP 連線失敗，請檢查帳號密碼: {e}")
        return

    # 開始遍歷名單 (假設從第 2 列開始是資料，視你的 Excel 而定)
    start_row = 2 
    
    for i in range(start_row, max_row + 1):
        # 讀取參加者資訊
        name = worksheet.cell(row=i, column=1).value
        email = worksheet.cell(row=i, column=4).value
        my_gift = worksheet.cell(row=i, column=5).value
        
        if not email:
            continue

        print(f"正在處理第 {i} 筆：{name} ({email})")

        # 建構信件內容
        content = [
            f"{name} 同學您好，\n",
            "由於逢期中考週，人手不足，本次活動有所延宕，在這裡說聲抱歉。",
            "配對結果目前正分批寄出。這裡為您複習提供的交換禮物：\n",
            f"【您提供的禮物】：\n{my_gift}\n",
            "--------------------------------------",
            "【您配對到的禮物】：\n"
        ]

        # 讀取配對對象 (假設配對 ID 放在第 6 欄之後)
        pairing_col_index = 6
        has_match = False
        
        while True:
            # 取得配對對象的 Row Index (假設 Excel 裡面存的是 Excel 的 Row ID)
            match_row_id = worksheet.cell(row=i, column=pairing_col_index).value
            
            if match_row_id is None:
                break
                
            match_row_id = int(match_row_id)
            has_match = True
            
            # 讀取對方資訊
            partner_name = worksheet.cell(row=match_row_id, column=1).value
            partner_gift = worksheet.cell(row=match_row_id, column=5).value
            partner_contact = worksheet.cell(row=match_row_id, column=4).value # 對方 Email
            
            content.append(f"對方姓名：{partner_name}")
            content.append(f"對方聯絡方式：{partner_contact}")
            content.append(f"禮物內容：\n{partner_gift}\n")
            content.append("----")
            
            pairing_col_index += 1

        content.append("\n回饋表單連結：[Link]\n謝謝您的參與！")
        
        # 寄出信件
        if has_match:
            try:
                subject = f"【非物質交換禮物】配對結果通知 - {name}"
                yag.send(to=email, subject=subject, contents=content)
                print(f"--> 已寄送給 {name}")
                
                # 避免觸發 Gmail 濫發信件機制，每寄一封休息一下
                time.sleep(2) 
            except Exception as e:
                print(f"--> 寄送失敗 {name}: {e}")
        else:
            print(f"--> {name} 沒有配對對象，跳過。")

if __name__ == "__main__":
    send_gift_emails()